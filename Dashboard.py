'''
NAME
    SquizrunAnalytics

Description
    Read an automated email via IMAP and get daily data of User Traffic, Game Statistics, and Survival
    Append/update the data to the currently used excel file
    Take user input and display the data on web accordingly using streamlit and plotly
'''
import configparser
import imaplib
import email
import openpyxl

import plotly.express as px 
import streamlit as st
import pandas as pd
import datetime


# list of indicators to display in each excel sheets "User Traffic" and "Game Statistics"
USER_TRAFFIC_INDIC = ['날짜', 'DAU(login)', 'DAU(chat)', 'DAU(quiz)', 'WAU(login)', 'MAU(login)',
                      'D+1 retention\n(login)', 'D+1 retention\n(nru login)', 'W+1 retention\n(login)',
                      'W+1 retention\n(new users login)', 'NRU (일일 신규 사용자 수)', 'CRU (누적 사용자 수)']
GAME_STATISTICS_INDIC = ['날짜', '스퀴즈볼 총 획득량', '스퀴즈볼 총 사용량\n(소모성아이템 한정)', 'QAR\n퀴즈 정답률', '총 퀴즈 참여자 수', 'NQPP\n인당 퀴즈 참여',
                        'NCPP\n인당 채팅 참여', 'IUPP\n인당 아이템 사용']


user_traffic = {}
game_statistics = {}


def email_fetch_append():
    ''' Read the the most recent gmail to get raw data for User Traffic, Game Statistics, and Survival, then update the current excel file.

    Parameters: 
        None
    Returns: 
        None
    '''

    excel_file = "2023_스퀴즈런_사용자지표_.xlsx"
    workbook = openpyxl.load_workbook(excel_file)

    # Get credentials from config file
    config = configparser.ConfigParser()
    config.read("config.ini")
    address = config.get("Credentials", "address")
    password = config.get("Credentials", "password")

    # Log in to Gmail
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(address, password)
    mail.select("inbox")
    print("LOGIN SUCCESSFUL")

    # Search for emails from specific address
    status, mail_id = mail.search(None, "FROM 'daily_report_squizrun@a.sktelecom.com'")

    # List of IDs of all the emails fetched
    mail_id_list = mail_id[0].split()

    # Fetch the email content
    latest_email_id = mail_id_list[-1]
    status, msg_data = mail.fetch(latest_email_id, "(RFC822)")
    print(f'latest_email_id : {latest_email_id}\n')

    # Parse the raw email
    msg = email.message_from_bytes(msg_data[0][1])

    email_body = (msg.get_payload(decode=True).decode()).rstrip()

    lines = email_body.split("\n")

    data_sets = {'user_traffic': [], 'game_statistics': [], 'survivor': []}

    # Group lines by data set
    for line in lines:
        indic_type = line.split(':')[0]
        data_sets[indic_type].append(line)

    # Sort lines within each data set by date
    for indic_type, data_lines in data_sets.items():
        data_sets[indic_type] = sorted(data_lines, key=lambda x: datetime.datetime.strptime(x.split(':')[1].split(',')[0], '%Y-%m-%d'))

    # Combine sorted lines into the final list
    sorted_lines = []
    for indic_type in ['user_traffic', 'game_statistics', 'survivor']:
        sorted_lines.extend(data_sets[indic_type])

    print("From       : {}".format(msg.get("From")))
    print("To         : {}".format(msg.get("To")))
    print("Bcc        : {}".format(msg.get("Bcc")))
    print("Date       : {}".format(msg.get("Date")))
    print("Subject    : {}\n".format(msg.get("Subject")))

    for line in sorted_lines:
        temp_split = line.split(":")
        sheet_data = temp_split[0]
        date_data = datetime.datetime.strptime((temp_split[1].split(","))[0], "%Y-%m-%d")
        num_data_str = ((temp_split[1]).split(","))[1:]
        
        # change data in num_data_str to appropriate data types
        num_data = []
        for item in num_data_str:
            if isinstance(item, int) or isinstance(item, float):
                num_data.append(item)  # Keep integers and floats as they are
            else:
                try:
                    num_data.append(int(item))  # Try converting to int
                except ValueError:
                    try:
                        num_data.append(float(item))  # Try converting to float
                    except ValueError:
                        num_data.append(item)  # If neither int nor float, keep as string

        print(f'sheet_data : {sheet_data}')
        print(f'date_data  : {date_data}')
        print(f'num_data   : {num_data}\n')

        if sheet_data == 'user_traffic':
            sheet = workbook['User Traffic']
        if sheet_data == 'game_statistics':
            sheet = workbook['Game Statistics']     
        if sheet_data == 'survivor':
            sheet = workbook['서바이벌 모드']   

        # find column with corresponding date
        target_row = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            if row[0].value == date_data:
                target_row = row[0].row
                break
        
        # find merged cell range
        merged_cells = sheet.merged_cells.ranges

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            # Check if the date in the first column matches date_data
            if row[0].value == date_data:
                data_index = 0
                for cell in row[1:]:
                    # If the cell is part of a merged range, find the initial cell
                    merged_range = None
                    for merged_range in merged_cells:
                        if cell.coordinate in merged_range:
                            initial_cell = merged_range.start_cell
                            break
                    else:
                        initial_cell = cell
                    if data_index < len(num_data):
                        # Write the value to the cell and move to the next value
                        initial_cell.value = num_data[data_index]
                        data_index += 1


    # Save the changes
    workbook.save(excel_file)
    workbook.close()

    print("EXCEL UPDATE COMPLETED")

    # Logout from Gmail
    mail.logout()


def load_from_file(file_name='2023_스퀴즈런_사용자지표_.xlsx') -> (pd.DataFrame, pd.DataFrame):
    ''' Read the excel files to create a DataFrame, and format the '날짜' column to show "yyyy-mm-dd day of the week"

    Parameters: 
        String: File name to read
    Returns: 
        DataFrame: 2D table of rows and columns
    '''
    df_ut = pd.read_excel(file_name, skiprows=range(3), sheet_name="User Traffic")
    df_gs = pd.read_excel(file_name, skiprows=range(2), sheet_name="Game Statistics")

    return df_ut, df_gs


def transform(df, indic):
    ''' Create dictionaries with indicator as key and a list of all its values as dictionary value

    Parameters: 
        DataFrame: 2D table of rows and columns
        List: list of indicators
    Returns: 
        Dictionary: indicator paired with list of its values
    '''
    data = {}
    for inx, value in enumerate(indic):
        data[value] = df[indic[inx]].tolist()

    return data


def draw(df: pd.DataFrame, indicator, height=800):
    '''
    Get DataFrame and an indicator string as parameters to create and display a line graph of that indicator 

    Parameters: 
        DataFrame: 2D table of rows and columns
        String: desired indicator of graph to create
    Returns: 
        Figure: object created with Plotly 
    '''
    
    if isinstance(indicator, str):
        indicator = [indicator]

    title = '   '.join(indicator)

    figure = px.line(df, x='날짜', y=indicator, text=None, markers=True, title=title, height=height)

    figure.update_traces(connectgaps=True)

    st.plotly_chart(figure,use_container_width=True)
    
    return figure


def draw_week(df_ut: pd.DataFrame, df_gs: pd.DataFrame):
    '''
    Get DataFrames as parameters to create and display bar graphs containing key weekly data 

    Parameters: 
        DataFrame: 2D table of rows and columns (User Traffic)
        DataFrame: 2D table of rows and columns (Game Statistics)
    Returns: 
        None 
    '''

    col1_1, col1_2, col1_3, col1_4 = st.columns([0.25, 0.25, 0.25, 0.25])

    end_date = datetime.date.today()
    start_date = end_date - datetime.timedelta(days=7)

    # Convert the date to a datetime object
    start_datetime = datetime.datetime.combine(start_date, datetime.datetime.min.time())
    end_datetime = datetime.datetime.combine(end_date, datetime.datetime.min.time())
    
    # filtered
    filtered_df_ut = df_ut[(df_ut['날짜'] >= start_datetime) & (df_ut['날짜'] <= end_datetime)]
    filtered_df_gs = df_gs[(df_gs['날짜'] >= start_datetime) & (df_gs['날짜'] <= end_datetime)]

    with col1_1:
        fig1 = px.bar(filtered_df_ut, x='날짜', y='DAU(login)', title='DAILY ACTIVE USERS (login)', width=300, height=300)
        fig1.update_layout(yaxis={'visible':False, 'showticklabels': False}, xaxis={'showticklabels': False}, title={'y':0.8}, yaxis_title=None, xaxis_title=None)
        st.plotly_chart(fig1)

    with col1_2:
        fig2 = px.bar(filtered_df_ut, x='날짜', y='NRU (일일 신규 사용자 수)', title='NRU', width=300, height=300)
        fig2.update_layout(yaxis={'visible':False, 'showticklabels': False}, xaxis={'showticklabels': False}, title={'y':0.8}, yaxis_title=None, xaxis_title=None)
        st.plotly_chart(fig2)

    with col1_3:
        fig3 = px.bar(filtered_df_ut, x='날짜', y='D+1 retention\n(login)', title='Rentetion (???)', width=300, height=300)
        fig3.update_layout(yaxis={'visible':False, 'showticklabels': False}, xaxis={'showticklabels': False}, title={'y':0.8}, yaxis_title=None, xaxis_title=None)
        st.plotly_chart(fig3)

    with col1_4:
        fig4 = px.bar(filtered_df_gs, x='날짜', y='IUPP\n인당 아이템 사용', title='IUPP', width=300, height=300)
        fig4.update_layout(yaxis={'visible':False, 'showticklabels': False}, xaxis={'showticklabels': False}, title={'y':0.8}, yaxis_title=None, xaxis_title=None)
        st.plotly_chart(fig4)



def date_range(df):
    '''
    Get user's choice of start date and end date to create a date range to return DataFrame within that date range

    Parameters: 
        DataFrame: 2D table of rows and columns
    Returns: 
        DataFrame: filtered Dataframe containing data within the chosen date range
    '''

    # Display the date_input widget 
    date_range = st.date_input(":date: Select Date Range", [df['날짜'].min().date(), datetime.date.today()], 
                               min_value=df['날짜'].min().date(), max_value=datetime.date.today())

    try:
        # Convert selected_date_range values to datetime objects
        start_date = datetime.datetime.combine(date_range[0], datetime.datetime.min.time())
        end_date = datetime.datetime.combine(date_range[1], datetime.datetime.min.time())
        filtered_df = df[(df['날짜'] >= start_date) & (df['날짜'] <= end_date)]
    except ValueError:
        # Proceed when two dates are chosen
        st.error("Please specify the end date")
        st.stop()
    except IndexError:
        st.error("Please specify the end date")
        st.stop()

    return filtered_df


def main():

    df_ut, df_gs = load_from_file()

    email_fetch_append()

    st.set_page_config(layout="wide")
    st.title("SQUIZRUN ANALYTICS")
    draw_week(df_ut, df_gs)

    col2_1, col2_2 = st.columns([0.17, 0.83])

    with col2_1:
        filtered_df = date_range(df_ut)
    draw(filtered_df, ['DAU(login)', 'MAU(login)', 'WAU(login)'], 500)
    
    

if __name__ == '__main__':
    main()
