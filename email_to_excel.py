import configparser
import imaplib
import email
import openpyxl
import datetime

def email_fetch_append():
    ''' Read the the most recent gmail to get raw data for User Traffic, Game Statistics, and Survival, then update the current excel file.

    Parameters: 
        None
    Returns: 
        None
    '''

    excel_file = "2023_스퀴즈런_사용자지표_.xlsx"
    workbook = openpyxl.load_workbook(excel_file)

    # Get credentials from config file
    config = configparser.ConfigParser()
    config.read("config.ini")
    address = config.get("Credentials", "address")
    password = config.get("Credentials", "password")

    # Log in to Gmail
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(address, password)
    mail.select("inbox")
    print("LOGIN SUCCESSFUL")

    # Search for emails from specific address
    status, mail_id = mail.search(None, "FROM 'daily_report_squizrun@a.sktelecom.com'")

    # List of IDs of all the emails fetched
    mail_id_list = mail_id[0].split()

    # Fetch the email content
    latest_email_id = mail_id_list[-1]
    status, msg_data = mail.fetch(latest_email_id, "(RFC822)")
    print(f'latest_email_id : {latest_email_id}\n')

    # Parse the raw email
    msg = email.message_from_bytes(msg_data[0][1])

    email_body = (msg.get_payload(decode=True).decode()).rstrip()

    lines = email_body.split("\n")

    data_sets = {'user_traffic': [], 'game_statistics': [], 'survivor': []}

    # Group lines by data set
    for line in lines:
        indic_type = line.split(':')[0]
        data_sets[indic_type].append(line)

    # Sort lines within each data set by date
    for indic_type, data_lines in data_sets.items():
        data_sets[indic_type] = sorted(data_lines, key=lambda x: datetime.datetime.strptime(x.split(':')[1].split(',')[0], '%Y-%m-%d'))

    # Combine sorted lines into the final list
    sorted_lines = []
    for indic_type in ['user_traffic', 'game_statistics', 'survivor']:
        sorted_lines.extend(data_sets[indic_type])

    print("From       : {}".format(msg.get("From")))
    print("To         : {}".format(msg.get("To")))
    print("Bcc        : {}".format(msg.get("Bcc")))
    print("Date       : {}".format(msg.get("Date")))
    print("Subject    : {}\n".format(msg.get("Subject")))

    for line in sorted_lines:
        temp_split = line.split(":")
        sheet_data = temp_split[0]
        date_data = datetime.datetime.strptime((temp_split[1].split(","))[0], "%Y-%m-%d")
        num_data_str = ((temp_split[1]).split(","))[1:]
        
        # change data in num_data_str to appropriate data types
        num_data = []
        for item in num_data_str:
            if isinstance(item, int) or isinstance(item, float):
                num_data.append(item)  # Keep integers and floats as they are
            else:
                try:
                    num_data.append(int(item))  # Try converting to int
                except ValueError:
                    try:
                        num_data.append(float(item))  # Try converting to float
                    except ValueError:
                        num_data.append(item)  # If neither int nor float, keep as string

        print(f'sheet_data : {sheet_data}')
        print(f'date_data  : {date_data}')
        print(f'num_data   : {num_data}\n')

        if sheet_data == 'user_traffic':
            sheet = workbook['User Traffic']
        if sheet_data == 'game_statistics':
            sheet = workbook['Game Statistics']     
        if sheet_data == 'survivor':
            sheet = workbook['서바이벌 모드']   

        # find column with corresponding date
        target_row = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            if row[0].value == date_data:
                target_row = row[0].row
                break
        
        # find merged cell range
        merged_cells = sheet.merged_cells.ranges

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            # Check if the date in the first column matches date_data
            if row[0].value == date_data:
                data_index = 0
                for cell in row[1:]:
                    # If the cell is part of a merged range, find the initial cell
                    merged_range = None
                    for merged_range in merged_cells:
                        if cell.coordinate in merged_range:
                            initial_cell = merged_range.start_cell
                            break
                    else:
                        initial_cell = cell
                    if data_index < len(num_data):
                        # Write the value to the cell and move to the next value
                        initial_cell.value = num_data[data_index]
                        data_index += 1


    # Save the changes
    workbook.save(excel_file)
    workbook.close()

    print("EXCEL UPDATE COMPLETED")

    # Logout from Gmail
    mail.logout()


def main():
    email_fetch_append()
    

if __name__ == '__main__':
    main()
